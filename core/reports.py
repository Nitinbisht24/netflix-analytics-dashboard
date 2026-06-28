# -*- coding: utf-8 -*-
"""
core/reports.py — On-demand executive report export.

Two formats, both built fresh from the same aggregate tables the
dashboard itself reads from (no separate "report data" path to drift
out of sync):
  - PDF    via reportlab (Platypus) — a short executive summary
  - Excel  via openpyxl — a multi-sheet workbook for further analysis

Both write to an in-memory BytesIO buffer; the Flask route streams it
straight back, nothing touches disk.
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak)

from core import analytics as an
from core import engagement as eng

NETFLIX_RED = colors.HexColor("#E50914")
DARK = colors.HexColor("#141414")
GREY = colors.HexColor("#564d4d")


def _country_label(country: str) -> str:
    return "Global Market" if not country or country == "All" else country


# ─────────────────────────────────────────────────────────────────────────
# PDF
# ─────────────────────────────────────────────────────────────────────────

def build_pdf_report(country: str) -> io.BytesIO:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.6 * inch, bottomMargin=0.6 * inch,
                             leftMargin=0.7 * inch, rightMargin=0.7 * inch)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleRed", parent=styles["Title"], textColor=NETFLIX_RED, fontSize=22)
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], textColor=DARK, spaceBefore=14, spaceAfter=6)
    body = ParagraphStyle("BodyGrey", parent=styles["Normal"], textColor=DARK, fontSize=10, leading=14)
    caption = ParagraphStyle("Caption", parent=styles["Normal"], textColor=GREY, fontSize=8, leading=11)

    label = _country_label(country)
    kpis = an.get_kpis(country)
    insights = an.get_insights(country)
    genres = an.get_roi_by_genre(country)
    top_titles = an.get_top_titles(country, limit=10)

    story = [
        Paragraph("Netflix Content Strategy Analytics", title_style),
        Paragraph(f"Executive Summary &mdash; {label}", styles["Heading3"]),
        Paragraph(f"Generated {datetime.now().strftime('%B %d, %Y')}", caption),
        Spacer(1, 16),
    ]

    kpi_data = [
        ["Titles Analyzed", f"{kpis['total']:,}"],
        ["Average Rating", f"{kpis['avg_rating']} / 10"],
        ["Average Budget", f"${kpis['avg_budget']:,}"],
        ["Average Revenue", f"${kpis['avg_revenue']:,}"],
        ["Market-Wide ROI", f"{kpis['roi']}%"],
        ["Leading Language", kpis["top_language"].upper()],
        ["Titles w/ Financial Data", f"{kpis['n_with_financials']:,} ({kpis['pct_with_financials']}%)"],
    ]
    t = Table(kpi_data, colWidths=[2.6 * inch, 3.4 * inch])
    t.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("TEXTCOLOR", (0, 0), (0, -1), GREY),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, -2), 0.5, colors.HexColor("#e0e0e0")),
    ]))
    story += [Paragraph("Key Performance Indicators", h2), t, Spacer(1, 10)]

    story.append(Paragraph("Top Insights", h2))
    for ins in insights[:6]:
        clean = ins["body"].replace("<b>", "<b>").replace("</b>", "</b>")
        story.append(Paragraph(f"&bull; {clean}", body))
        story.append(Spacer(1, 4))

    story.append(PageBreak())
    story.append(Paragraph(f"ROI by Genre &mdash; {label}", h2))
    genre_rows = [["Genre", "Titles", "Avg ROI", "Avg Rating", "Avg Budget ($M)", "Avg Revenue ($M)"]]
    for i in range(len(genres["genres"])):
        genre_rows.append([
            genres["genres"][i], genres["counts"][i],
            f"{genres['roi'][i]}%" if genres["roi"][i] is not None else "\u2014",
            genres["ratings"][i], f"${genres['budgets'][i]}M", f"${genres['revenues'][i]}M",
        ])
    gt = Table(genre_rows, colWidths=[1.3 * inch] * 6)
    gt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), DARK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dddddd")),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
    ]))
    story += [gt, Spacer(1, 16)]

    story.append(Paragraph(f"Top 10 Titles by Popularity &mdash; {label}", h2))
    title_rows = [["Title", "Year", "Rating", "Revenue ($M)", "ROI"]]
    for t_ in top_titles["titles"]:
        title_rows.append([
            t_["title"][:38], t_["release_year"], t_["vote_average"],
            f"${t_['revenue']}M", f"{t_['roi']}%" if t_["roi"] is not None else "\u2014",
        ])
    tt = Table(title_rows, colWidths=[2.6 * inch, 0.7 * inch, 0.8 * inch, 1.1 * inch, 0.8 * inch])
    tt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NETFLIX_RED),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dddddd")),
    ]))
    story += [tt, Spacer(1, 14)]

    story.append(Paragraph(
        "Data sources: global movie-market metadata (budget/revenue economics, used as a market "
        "benchmark) and Netflix's officially published bi-annual Engagement Reports. "
        "See README / DATA_DICTIONARY for full methodology.", caption))

    doc.build(story)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────
# Excel
# ─────────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="E50914", end_color="E50914")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="141414")


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_excel_report(country: str) -> io.BytesIO:
    label = _country_label(country)
    wb = Workbook()

    # ---- Sheet 1: KPIs ----
    ws = wb.active
    ws.title = "Summary"
    kpis = an.get_kpis(country)
    ws["A1"] = "Netflix Content Strategy Analytics"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Scope: {label}    Generated: {datetime.now().strftime('%Y-%m-%d')}"
    rows = [
        ("Metric", "Value"),
        ("Titles Analyzed", kpis["total"]),
        ("Average Rating (/10)", kpis["avg_rating"]),
        ("Average Budget ($)", kpis["avg_budget"]),
        ("Average Revenue ($)", kpis["avg_revenue"]),
        ("Total Budget ($)", kpis["total_budget"]),
        ("Total Revenue ($)", kpis["total_revenue"]),
        ("Market-Wide ROI (%)", kpis["roi"]),
        ("Leading Language", kpis["top_language"]),
        ("Titles w/ Financial Data", kpis["n_with_financials"]),
        ("% w/ Financial Data", kpis["pct_with_financials"]),
        ("YoY Growth (%)", kpis["yoy_growth"]),
    ]
    for r, (k, v) in enumerate(rows, start=4):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
    _style_header(ws, row=4)
    _autosize(ws, [28, 18])

    # ---- Sheet 2: Genre ROI ----
    ws2 = wb.create_sheet("Genre ROI")
    g = an.get_roi_by_genre(country)
    ws2.append(["Genre", "Title Count", "Avg ROI (%)", "Avg Rating", "Avg Budget ($M)", "Avg Revenue ($M)"])
    for i in range(len(g["genres"])):
        ws2.append([g["genres"][i], g["counts"][i], g["roi"][i], g["ratings"][i], g["budgets"][i], g["revenues"][i]])
    _style_header(ws2)
    _autosize(ws2, [18, 12, 12, 12, 16, 16])
    if len(g["genres"]) > 1:
        chart = BarChart()
        chart.title = "Average ROI by Genre (%)"
        chart.y_axis.title = "ROI (%)"
        data = Reference(ws2, min_col=3, min_row=1, max_row=len(g["genres"]) + 1)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=len(g["genres"]) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws2.add_chart(chart, "H2")

    # ---- Sheet 3: Top Titles ----
    ws3 = wb.create_sheet("Top Titles")
    titles = an.get_top_titles(country, limit=50)["titles"]
    ws3.append(["Title", "Year", "Rating", "Popularity", "Budget ($M)", "Revenue ($M)", "ROI (%)", "Genres", "Language"])
    for t_ in titles:
        ws3.append([t_["title"], t_["release_year"], t_["vote_average"], t_["popularity"],
                    t_["budget"], t_["revenue"], t_["roi"], t_["genres"], t_["language"]])
    _style_header(ws3)
    _autosize(ws3, [38, 8, 9, 11, 12, 13, 10, 30, 10])

    # ---- Sheet 4: Real Netflix Engagement ----
    ws4 = wb.create_sheet("Netflix Engagement")
    top_eng = eng.get_engagement_top("views", limit=50)["titles"]
    ws4.append(["Title", "Type", "Genre", "Country", "Report Period", "Hours Viewed (M)", "Views (M)"])
    for r in top_eng:
        ws4.append([r["title"], r["content_type"], r["primary_genre"], r["country_origin"],
                    r["report_period"], r["hours_viewed_millions"], r["views_millions"]])
    _style_header(ws4)
    _autosize(ws4, [42, 12, 14, 16, 14, 16, 12])
    ws4.append([])
    ws4.append(["Source: Netflix official 'What We Watched' Engagement Reports (H1 2023 \u2013 H2 2025) "
                "and Wikipedia's list of most-watched Netflix programming."])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
